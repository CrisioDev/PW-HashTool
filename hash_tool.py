#!/usr/bin/env python3
"""
Hash Generator Tool
Liest Excel/CSV-Dateien, generiert NTLM- oder SHA1-Hashes für ausgewählte Spalten
und maskiert die Originaldaten.
"""

from __future__ import annotations

import csv
import hashlib
import os
import struct
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Any, Callable

# Versuche openpyxl zu importieren (für Excel)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Konstanten
PASSWORD_KEYWORDS = ("password", "passwort", "pwd", "kennwort", "pass")
SUPPORTED_DELIMITERS = [";", ",", "\t", "|"]
SUPPORTED_ENCODINGS = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
CSV_SNIFF_SAMPLE_SIZE = 8192
MAX_COMBINED_FILENAME_LENGTH = 100
MAX_COLUMN_WIDTH = 50

# MD4 Runden-Konfiguration
MD4_ROUND2_ORDER = (0, 4, 8, 12, 1, 5, 9, 13, 2, 6, 10, 14, 3, 7, 11, 15)
MD4_ROUND3_ORDER = (0, 8, 4, 12, 2, 10, 6, 14, 1, 9, 5, 13, 3, 11, 7, 15)
MD4_ROUND2_SHIFTS = (3, 5, 9, 13)
MD4_ROUND3_SHIFTS = (3, 9, 11, 15)


def column_index_to_letter(index: int) -> str:
    """Wandelt einen Spaltenindex (1-basiert) in Excel-Buchstaben um."""
    result = ""
    while index > 0:
        index -= 1
        result = chr(65 + (index % 26)) + result
        index //= 26
    return result


def generate_ntlm_hash(text: str) -> str:
    """Generiert einen NTLM-Hash (MD4 auf UTF-16LE)."""
    if not text:
        return ""
    try:
        encoded = text.encode('utf-16-le')
        md4_hash = hashlib.new('md4', encoded)
        return md4_hash.hexdigest().upper()
    except ValueError:
        return md4_alternative(text.encode('utf-16-le'))


def generate_sha1_hash(text: str) -> str:
    """Generiert einen SHA1-Hash."""
    if not text:
        return ""
    encoded = text.encode('utf-8')
    sha1_hash = hashlib.sha1(encoded)
    return sha1_hash.hexdigest().upper()


def md4_alternative(data: bytes) -> str:
    """Alternative MD4-Implementierung falls hashlib MD4 nicht unterstützt."""

    def left_rotate(x: int, n: int) -> int:
        return ((x << n) | (x >> (32 - n))) & 0xFFFFFFFF

    def f(x: int, y: int, z: int) -> int:
        return (x & y) | (~x & z)

    def g(x: int, y: int, z: int) -> int:
        return (x & y) | (x & z) | (y & z)

    def h(x: int, y: int, z: int) -> int:
        return x ^ y ^ z

    # Initialisierungswerte
    a0 = 0x67452301
    b0 = 0xEFCDAB89
    c0 = 0x98BADCFE
    d0 = 0x10325476

    # Padding
    msg = bytearray(data)
    msg_len = len(data)
    msg.append(0x80)

    while (len(msg) % 64) != 56:
        msg.append(0x00)

    msg += struct.pack('<Q', msg_len * 8)

    # Verarbeite in 64-Byte Blöcken
    for i in range(0, len(msg), 64):
        block = msg[i:i+64]
        X = list(struct.unpack('<16I', block))

        A, B, C, D = a0, b0, c0, d0

        # Runde 1
        for k in range(16):
            if k % 4 == 0:
                A = left_rotate((A + f(B, C, D) + X[k]) & 0xFFFFFFFF, 3)
            elif k % 4 == 1:
                D = left_rotate((D + f(A, B, C) + X[k]) & 0xFFFFFFFF, 7)
            elif k % 4 == 2:
                C = left_rotate((C + f(D, A, B) + X[k]) & 0xFFFFFFFF, 11)
            else:
                B = left_rotate((B + f(C, D, A) + X[k]) & 0xFFFFFFFF, 19)

        # Runde 2
        for idx, k in enumerate(MD4_ROUND2_ORDER):
            shift = MD4_ROUND2_SHIFTS[idx % 4]
            if idx % 4 == 0:
                A = left_rotate((A + g(B, C, D) + X[k] + 0x5A827999) & 0xFFFFFFFF, shift)
            elif idx % 4 == 1:
                D = left_rotate((D + g(A, B, C) + X[k] + 0x5A827999) & 0xFFFFFFFF, shift)
            elif idx % 4 == 2:
                C = left_rotate((C + g(D, A, B) + X[k] + 0x5A827999) & 0xFFFFFFFF, shift)
            else:
                B = left_rotate((B + g(C, D, A) + X[k] + 0x5A827999) & 0xFFFFFFFF, shift)

        # Runde 3
        for idx, k in enumerate(MD4_ROUND3_ORDER):
            shift = MD4_ROUND3_SHIFTS[idx % 4]
            if idx % 4 == 0:
                A = left_rotate((A + h(B, C, D) + X[k] + 0x6ED9EBA1) & 0xFFFFFFFF, shift)
            elif idx % 4 == 1:
                D = left_rotate((D + h(A, B, C) + X[k] + 0x6ED9EBA1) & 0xFFFFFFFF, shift)
            elif idx % 4 == 2:
                C = left_rotate((C + h(D, A, B) + X[k] + 0x6ED9EBA1) & 0xFFFFFFFF, shift)
            else:
                B = left_rotate((B + h(C, D, A) + X[k] + 0x6ED9EBA1) & 0xFFFFFFFF, shift)

        a0 = (a0 + A) & 0xFFFFFFFF
        b0 = (b0 + B) & 0xFFFFFFFF
        c0 = (c0 + C) & 0xFFFFFFFF
        d0 = (d0 + D) & 0xFFFFFFFF

    return struct.pack('<4I', a0, b0, c0, d0).hex().upper()


def mask_value(text: str) -> str:
    """Maskiert den Wert: erste 3 Zeichen + ***"""
    if not text:
        return ""
    if len(text) <= 3:
        return text[0] + "***" if len(text) >= 1 else "***"
    return text[:3] + "***"


class HashTool:
    def __init__(self, root: tk.Tk) -> None:
        self.root: tk.Tk = root
        self.root.title("Hash Generator Tool")
        self.root.geometry("650x500")
        self.root.resizable(True, True)

        self.file_path: str | None = None
        self.file_paths: list[str] = []
        self.columns: list[tuple[int, str]] = []
        self.file_type: str | None = None
        self.workbook: Any = None
        self.csv_data: list[list[str]] | None = None

        self.create_widgets()

    def create_widgets(self) -> None:
        """Erstellt die GUI-Elemente."""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        self._create_file_section(main_frame)
        self._create_options_section(main_frame)
        self._create_action_section(main_frame)

        self.log("Bereit. Bitte eine Excel- oder CSV-Datei auswählen.")
        if not EXCEL_SUPPORT:
            self.log("HINWEIS: openpyxl nicht installiert. Excel-Support deaktiviert.")
            self.log("Installieren mit: pip install openpyxl")

    def _create_file_section(self, parent: ttk.Frame) -> None:
        """Erstellt den Datei-Auswahl-Bereich."""
        ttk.Label(parent, text="Datei:").grid(row=0, column=0, sticky="w", pady=5)

        file_frame = ttk.Frame(parent)
        file_frame.grid(row=0, column=1, sticky="ew", pady=5)
        file_frame.columnconfigure(0, weight=1)

        self.file_entry = ttk.Entry(file_frame)
        self.file_entry.grid(row=0, column=0, sticky="ew", padx=(0, 5))

        ttk.Button(file_frame, text="Durchsuchen...", command=self.browse_file).grid(row=0, column=1)
        ttk.Button(file_frame, text="\u21BB", width=3, command=self.load_file).grid(row=0, column=2, padx=(5, 0))

    def _create_options_section(self, parent: ttk.Frame) -> None:
        """Erstellt Header-, Sheet-, Spalten-, Hash-Typ- und CSV-Optionen."""
        # Header-Option
        self.has_header_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            parent,
            text="Erste Zeile ist Header/Titel (nicht hashen)",
            variable=self.has_header_var,
            command=self.on_header_option_changed
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=5)

        # Excel Sheet Auswahl
        ttk.Label(parent, text="Sheet:").grid(row=2, column=0, sticky="w", pady=5)
        self.sheet_combo = ttk.Combobox(parent, state="readonly")
        self.sheet_combo.grid(row=2, column=1, sticky="ew", pady=5)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_selected)

        # Spalten-Auswahl
        ttk.Label(parent, text="Spalte:").grid(row=3, column=0, sticky="w", pady=5)
        self.column_combo = ttk.Combobox(parent, state="readonly")
        self.column_combo.grid(row=3, column=1, sticky="ew", pady=5)

        # Hash-Typ Auswahl
        hash_frame = ttk.LabelFrame(parent, text="Hash-Typ", padding="5")
        hash_frame.grid(row=4, column=0, columnspan=2, sticky="ew", pady=10)

        self.hash_ntlm_var = tk.BooleanVar(value=True)
        self.hash_sha1_var = tk.BooleanVar(value=True)

        ttk.Checkbutton(hash_frame, text="NTLM (MD4, Windows-Passwörter)",
                        variable=self.hash_ntlm_var).pack(side="left", padx=10)
        ttk.Checkbutton(hash_frame, text="SHA1",
                        variable=self.hash_sha1_var).pack(side="left", padx=10)

        # CSV Optionen
        csv_frame = ttk.LabelFrame(parent, text="CSV-Optionen", padding="5")
        csv_frame.grid(row=5, column=0, columnspan=2, sticky="ew", pady=10)
        csv_frame.columnconfigure(1, weight=1)

        ttk.Label(csv_frame, text="Trennzeichen:").grid(row=0, column=0, sticky="w")
        self.delimiter_var = tk.StringVar(value=",")
        self.delimiter_combo = ttk.Combobox(csv_frame, textvariable=self.delimiter_var,
                                            values=SUPPORTED_DELIMITERS, width=5)
        self.delimiter_combo.grid(row=0, column=1, sticky="w", padx=5)

        ttk.Label(csv_frame, text="Encoding:").grid(row=0, column=2, sticky="w", padx=(20, 0))
        self.encoding_var = tk.StringVar(value="utf-8")
        self.encoding_combo = ttk.Combobox(csv_frame, textvariable=self.encoding_var,
                                           values=SUPPORTED_ENCODINGS, width=10)
        self.encoding_combo.grid(row=0, column=3, sticky="w", padx=5)

    def _create_action_section(self, parent: ttk.Frame) -> None:
        """Erstellt Buttons und Status/Log-Bereich."""
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=6, column=0, columnspan=2, pady=20)

        ttk.Button(button_frame, text="Verarbeiten & Speichern", command=self.process_file).pack(side="left", padx=5)

        # Status/Log
        ttk.Label(parent, text="Status:").grid(row=7, column=0, sticky="nw", pady=5)

        self.status_text = tk.Text(parent, height=8, width=50)
        self.status_text.grid(row=7, column=1, sticky="nsew", pady=5)
        parent.rowconfigure(7, weight=1)

        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.status_text.yview)
        scrollbar.grid(row=7, column=2, sticky="ns", pady=5)
        self.status_text.configure(yscrollcommand=scrollbar.set)

    def log(self, message: str) -> None:
        """Schreibt eine Nachricht in das Status-Fenster."""
        self.status_text.insert("end", message + "\n")
        self.status_text.see("end")

    def on_header_option_changed(self) -> None:
        """Wird aufgerufen wenn die Header-Checkbox geändert wird."""
        if self.file_path and self.file_type:
            self.update_column_list()

    def update_column_list(self) -> None:
        """Aktualisiert die Spaltenliste basierend auf Header-Option."""
        if self.file_type == 'excel':
            self.update_excel_columns()
        elif self.file_type == 'csv':
            self.update_csv_columns()

    def _populate_column_list(self, column_values: list) -> None:
        """Befüllt die Spalten-Combobox aus einer Liste von Spaltenwerten."""
        has_header = self.has_header_var.get()
        self.columns = []
        for i, cell_value in enumerate(column_values):
            col = i + 1
            col_letter = column_index_to_letter(col)
            if has_header and cell_value:
                display_name = f"{cell_value} (Spalte {col_letter})"
            else:
                display_name = f"Spalte {col_letter}"
            self.columns.append((col, display_name))

        self.column_combo['values'] = [name for _, name in self.columns]
        if self.columns:
            self.column_combo.current(0)
            self.auto_select_password_column()

    def update_excel_columns(self) -> None:
        """Aktualisiert Excel-Spalten basierend auf Header-Option."""
        sheet_name = self.sheet_combo.get()
        if not sheet_name or not self.workbook:
            return
        sheet = self.workbook[sheet_name]
        column_values = [
            sheet.cell(row=1, column=col).value
            for col in range(1, sheet.max_column + 1)
        ]
        self._populate_column_list(column_values)

    def update_csv_columns(self) -> None:
        """Aktualisiert CSV-Spalten basierend auf Header-Option."""
        if not self.csv_data or len(self.csv_data) == 0:
            return
        self._populate_column_list(self.csv_data[0])

    def auto_select_password_column(self) -> None:
        """Wählt automatisch die Passwort-Spalte vor, falls erkannt."""
        for i, (_, display_name) in enumerate(self.columns):
            name_lower = display_name.lower().split(' (spalte')[0].strip()
            if name_lower in PASSWORD_KEYWORDS:
                self.column_combo.current(i)
                self.log(f"Passwort-Spalte automatisch erkannt: {display_name}")
                return

    def browse_file(self) -> None:
        """Öffnet den Datei-Auswahl-Dialog."""
        filetypes = [("Alle unterstützten", "*.xlsx *.csv")]
        if EXCEL_SUPPORT:
            filetypes.append(("Excel-Dateien", "*.xlsx"))
        filetypes.append(("CSV-Dateien", "*.csv"))
        filetypes.append(("Alle Dateien", "*.*"))

        paths = filedialog.askopenfilenames(filetypes=filetypes)
        if paths:
            self.file_paths = list(paths)
            self.file_entry.delete(0, "end")
            if len(paths) == 1:
                self.file_entry.insert(0, paths[0])
            else:
                self.file_entry.insert(0, f"{len(paths)} Dateien ausgewählt")
            self.load_file()

    def load_file(self) -> None:
        """Lädt die ausgewählte(n) Datei(en)."""
        # Multi-CSV-Modus
        if len(self.file_paths) > 1:
            all_csv = all(os.path.splitext(p)[1].lower() == '.csv' for p in self.file_paths)
            if not all_csv:
                messagebox.showerror("Fehler", "Mehrfachauswahl ist nur für CSV-Dateien möglich.")
                return
            for p in self.file_paths:
                if not os.path.exists(p):
                    messagebox.showerror("Fehler", f"Datei nicht gefunden: {p}")
                    return
            try:
                self.load_multiple_csv_files(self.file_paths)
            except Exception as e:
                self.log(f"Fehler beim Laden: {str(e)}")
                messagebox.showerror("Fehler", f"Fehler beim Laden:\n{str(e)}")
            return

        # Einzeldatei-Modus
        path = self.file_paths[0] if self.file_paths else self.file_entry.get()
        if not path:
            messagebox.showerror("Fehler", "Bitte zuerst eine Datei auswählen.")
            return

        if not os.path.exists(path):
            messagebox.showerror("Fehler", f"Datei nicht gefunden: {path}")
            return

        self.file_path = path
        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == ".xlsx":
                if not EXCEL_SUPPORT:
                    messagebox.showerror("Fehler", "Excel-Support nicht verfügbar. Bitte openpyxl installieren.")
                    return
                self.load_excel_file(path)
            elif ext == ".csv":
                self.load_csv_file(path)
            else:
                messagebox.showerror("Fehler", "Nicht unterstütztes Dateiformat. Bitte .xlsx oder .csv verwenden.")
        except Exception as e:
            self.log(f"Fehler beim Laden: {str(e)}")
            messagebox.showerror("Fehler", f"Fehler beim Laden der Datei:\n{str(e)}")

    def load_excel_file(self, path: str) -> None:
        """Lädt eine Excel-Datei."""
        self.file_type = 'excel'
        self.workbook = load_workbook(path)
        self.csv_data = None

        self.sheet_combo['values'] = self.workbook.sheetnames
        self.sheet_combo.current(0)
        self.sheet_combo['state'] = 'readonly'

        self.on_sheet_selected(None)
        self.log(f"Excel-Datei geladen: {os.path.basename(path)}")
        self.log(f"Verfügbare Sheets: {', '.join(self.workbook.sheetnames)}")

    def on_sheet_selected(self, event: Any) -> None:
        """Wird aufgerufen wenn ein Sheet ausgewählt wird."""
        if self.file_type != 'excel':
            return
        sheet_name = self.sheet_combo.get()
        if not sheet_name:
            return
        self.update_excel_columns()
        self.log(f"Sheet '{sheet_name}' geladen. {len(self.columns)} Spalten gefunden.")

    def _read_csv_file(self, path: str) -> tuple[list[list[str]], str]:
        """Liest eine CSV-Datei mit automatischer Delimiter-Erkennung."""
        encoding = self.encoding_var.get()
        delimiter = self.detect_csv_delimiter(path, encoding)
        if delimiter == "\\t":
            delimiter = "\t"
        with open(path, 'r', encoding=encoding, newline='') as f:
            reader = csv.reader(f, delimiter=delimiter)
            return list(reader), delimiter

    def load_multiple_csv_files(self, paths: list[str]) -> None:
        """Lädt mehrere CSV-Dateien und fügt sie zusammen. Header müssen identisch sein."""
        self.file_type = 'csv'
        self.workbook = None
        self.sheet_combo['state'] = 'disabled'
        self.sheet_combo.set('')

        reference_header: list[str] | None = None
        all_data_rows: list[list[str]] = []

        for path in paths:
            try:
                rows, delimiter = self._read_csv_file(path)
            except UnicodeDecodeError:
                self.log(f"Encoding-Fehler bei: {os.path.basename(path)}")
                messagebox.showerror("Fehler", f"Encoding-Fehler bei:\n{os.path.basename(path)}")
                return

            if not rows:
                self.log(f"Übersprungen (leer): {os.path.basename(path)}")
                continue

            header = rows[0]
            data = rows[1:]

            if reference_header is None:
                reference_header = header
                self.delimiter_var.set(delimiter)
            else:
                if header != reference_header:
                    diff_file = os.path.basename(path)
                    ref_file = os.path.basename(paths[0])
                    messagebox.showerror(
                        "Fehler",
                        f"Spaltenstruktur unterschiedlich!\n\n"
                        f"'{diff_file}' stimmt nicht mit '{ref_file}' überein.\n\n"
                        f"Alle CSVs müssen identische Spalten haben."
                    )
                    return

            all_data_rows.extend(data)
            self.log(f"Geladen: {os.path.basename(path)} ({len(data)} Zeilen)")

        if reference_header is None:
            messagebox.showerror("Fehler", "Alle CSV-Dateien sind leer.")
            return

        self.csv_data = [reference_header] + all_data_rows
        self.file_path = paths[0]

        self.update_csv_columns()
        self.log(f"Zusammengeführt: {len(paths)} Dateien, {len(all_data_rows)} Datenzeilen")

    def detect_csv_delimiter(self, path: str, encoding: str) -> str:
        """Erkennt automatisch das Trennzeichen einer CSV-Datei."""
        try:
            with open(path, 'r', encoding=encoding, newline='') as f:
                sample = f.read(CSV_SNIFF_SAMPLE_SIZE)
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample, delimiters=',;\t|')
            return dialect.delimiter
        except csv.Error:
            return self.delimiter_var.get()

    def load_csv_file(self, path: str) -> None:
        """Lädt eine einzelne CSV-Datei."""
        self.file_type = 'csv'
        self.workbook = None
        self.sheet_combo['state'] = 'disabled'
        self.sheet_combo.set('')

        try:
            rows, delimiter = self._read_csv_file(path)
            self.delimiter_var.set(delimiter)
        except UnicodeDecodeError:
            encoding = self.encoding_var.get()
            self.log(f"Encoding-Fehler. Versuchen Sie ein anderes Encoding (aktuell: {encoding})")
            messagebox.showerror("Fehler", "Encoding-Fehler. Bitte anderes Encoding auswählen.")
            return

        if not rows:
            messagebox.showerror("Fehler", "CSV-Datei ist leer.")
            return

        self.csv_data = rows
        self.update_csv_columns()
        self.log(f"CSV-Datei geladen: {os.path.basename(path)}")
        self.log(f"Trennzeichen automatisch erkannt: '{delimiter}'")
        self.log(f"{len(self.columns)} Spalten gefunden.")

    def get_hash_functions(self) -> list[tuple[Callable[[str], str], str]]:
        """Gibt Liste der ausgewählten Hash-Funktionen zurück."""
        functions: list[tuple[Callable[[str], str], str]] = []
        if self.hash_ntlm_var.get():
            functions.append((generate_ntlm_hash, "NTLM"))
        if self.hash_sha1_var.get():
            functions.append((generate_sha1_hash, "SHA1"))
        return functions

    def process_file(self) -> None:
        """Verarbeitet die geladene Datei."""
        if not self.file_path or not self.columns:
            messagebox.showerror("Fehler", "Bitte zuerst eine Datei laden.")
            return

        selected = self.column_combo.get()
        if not selected:
            messagebox.showerror("Fehler", "Bitte eine Spalte auswählen.")
            return

        hash_functions = self.get_hash_functions()
        if not hash_functions:
            messagebox.showerror("Fehler", "Bitte mindestens einen Hash-Typ auswählen.")
            return

        selected_index = self.column_combo.current()
        if selected_index < 0 or selected_index >= len(self.columns):
            messagebox.showerror("Fehler", "Ungültige Spaltenauswahl.")
            return

        col_index = self.columns[selected_index][0]

        try:
            if self.file_type == 'excel':
                self.process_excel(col_index, hash_functions)
            else:
                self.process_csv(col_index, hash_functions)
        except Exception as e:
            self.log(f"Fehler bei Verarbeitung: {str(e)}")
            messagebox.showerror("Fehler", f"Fehler bei der Verarbeitung:\n{str(e)}")

    def process_excel(self, col_index: int, hash_functions: list[tuple]) -> None:
        """Verarbeitet eine Excel-Datei."""
        sheet_name = self.sheet_combo.get()
        sheet = self.workbook[sheet_name]
        has_header = self.has_header_var.get()

        hash_cols, processed_count = self._hash_excel_sheet(
            sheet, col_index, hash_functions, has_header
        )
        self._expand_excel_tables(sheet, len(hash_functions))
        self._save_excel_results(hash_cols, processed_count)

    def _hash_excel_sheet(
        self, sheet: Any, col_index: int, hash_functions: list[tuple], has_header: bool
    ) -> tuple[list[tuple], int]:
        """Fügt Hash-Spalten ein und hasht die Werte im Excel-Sheet."""
        col_letter = column_index_to_letter(col_index)
        original_header = sheet.cell(row=1, column=col_index).value or f"Spalte_{col_letter}"

        num_hashes = len(hash_functions)
        sheet.insert_cols(col_index + 1, amount=num_hashes)

        hash_cols = []
        for i, (hash_func, hash_name) in enumerate(hash_functions):
            hash_col = col_index + 1 + i
            hash_cols.append((hash_col, hash_func, hash_name))
            if has_header:
                sheet.cell(row=1, column=hash_col).value = f"{original_header}_{hash_name}_Hash"
            else:
                sheet.cell(row=1, column=hash_col).value = f"Spalte_{col_letter}_{hash_name}_Hash"

        start_row = 2 if has_header else 1
        processed_count = 0

        for row in range(start_row, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col_index)
            value = cell.value
            if value:
                value_str = str(value)
                for hash_col, hash_func, hash_name in hash_cols:
                    sheet.cell(row=row, column=hash_col).value = hash_func(value_str)
                cell.value = mask_value(value_str)
                processed_count += 1

        return hash_cols, processed_count

    def _expand_excel_tables(self, sheet: Any, num_hashes: int) -> None:
        """Erweitert Excel-Tabellen-Referenzen um die neuen Hash-Spalten."""
        for table in sheet.tables.values():
            ref = table.ref
            start, end = ref.split(':')
            end_col = ''.join(c for c in end if c.isalpha())
            end_row = ''.join(c for c in end if c.isdigit())
            old_end_col_idx = column_index_from_string(end_col)
            new_end_col_idx = old_end_col_idx + num_hashes
            new_end_col = get_column_letter(new_end_col_idx)
            table.ref = f"{start}:{new_end_col}{end_row}"

    def _save_excel_results(self, hash_cols: list[tuple], processed_count: int) -> None:
        """Speichert die verarbeitete Excel-Datei."""
        base, ext = os.path.splitext(self.file_path)
        output_path = f"{base}_hashed{ext}"

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel-Dateien", "*.xlsx")],
            initialfile=os.path.basename(output_path),
            initialdir=os.path.dirname(self.file_path)
        )

        if save_path:
            self.workbook.save(save_path)
            hash_names = ", ".join([name for _, _, name in hash_cols])
            self.log("Verarbeitung abgeschlossen!")
            self.log(f"{processed_count} Werte mit {hash_names} gehasht und maskiert.")
            self.log(f"Gespeichert unter: {save_path}")
            messagebox.showinfo("Erfolg", f"Datei erfolgreich gespeichert!\n\n{processed_count} Werte verarbeitet.")

    def process_csv(self, col_index: int, hash_functions: list[tuple]) -> None:
        """Verarbeitet eine CSV-Datei."""
        rows, processed_count = self._prepare_csv_rows(col_index, hash_functions)
        if rows is not None:
            self._save_csv_results(rows, hash_functions, processed_count)

    def _prepare_csv_rows(
        self, col_index: int, hash_functions: list[tuple]
    ) -> tuple[list[list[str]] | None, int]:
        """Bereitet CSV-Zeilen vor: fügt Hash-Spalten ein und hasht Werte."""
        has_header = self.has_header_var.get()

        # Arbeitskopie aus gespeicherten Daten
        rows = [row[:] for row in self.csv_data]

        if not rows:
            messagebox.showerror("Fehler", "CSV-Datei ist leer.")
            return None, 0

        col_idx = col_index - 1  # 0-basiert
        col_letter = column_index_to_letter(col_index)

        if has_header:
            header = rows[0]
            original_header = header[col_idx] if col_idx < len(header) else f"Spalte_{col_letter}"
            for hash_func, hash_name in reversed(hash_functions):
                header.insert(col_idx + 1, f"{original_header}_{hash_name}_Hash")
            start_row = 1
        else:
            new_header = [f"Spalte_{column_index_to_letter(i+1)}" for i in range(len(rows[0]))]
            for hash_func, hash_name in reversed(hash_functions):
                new_header.insert(col_idx + 1, f"Spalte_{col_letter}_{hash_name}_Hash")
            rows.insert(0, new_header)
            start_row = 1

        processed_count = 0
        num_hashes = len(hash_functions)

        for i in range(start_row, len(rows)):
            row = rows[i]
            if col_idx < len(row):
                value = row[col_idx]
                if value:
                    for hash_func, hash_name in reversed(hash_functions):
                        row.insert(col_idx + 1, hash_func(value))
                    row[col_idx] = mask_value(value)
                    processed_count += 1
                else:
                    for _ in range(num_hashes):
                        row.insert(col_idx + 1, "")
            else:
                for _ in range(num_hashes):
                    row.insert(col_idx + 1, "")

        return rows, processed_count

    def _save_csv_results(
        self, rows: list[list[str]], hash_functions: list[tuple], processed_count: int
    ) -> None:
        """Speichert CSV- und Excel-Ergebnis-Dateien."""
        delimiter = self.delimiter_var.get()
        if delimiter == "\\t":
            delimiter = "\t"
        encoding = self.encoding_var.get()

        if len(self.file_paths) > 1:
            names = [os.path.splitext(os.path.basename(p))[0] for p in self.file_paths]
            combined = "_".join(names)
            if len(combined) > MAX_COMBINED_FILENAME_LENGTH:
                combined = "_".join(names[:2]) + f"_+{len(names)-2}"
            output_name = f"{combined}_merged_hashed.csv"
        else:
            base, ext = os.path.splitext(self.file_path)
            output_name = os.path.basename(f"{base}_hashed{ext}")

        save_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV-Dateien", "*.csv")],
            initialfile=output_name,
            initialdir=os.path.dirname(self.file_path)
        )

        if save_path:
            with open(save_path, 'w', encoding=encoding, newline='') as f:
                writer = csv.writer(f, delimiter=delimiter)
                writer.writerows(rows)

            excel_path = os.path.splitext(save_path)[0] + ".xlsx"
            self.save_csv_as_excel(rows, excel_path)

            hash_names = ", ".join([name for _, name in hash_functions])
            self.log("Verarbeitung abgeschlossen!")
            self.log(f"{processed_count} Werte mit {hash_names} gehasht und maskiert.")
            self.log(f"CSV gespeichert unter: {save_path}")
            self.log(f"Excel gespeichert unter: {excel_path}")
            messagebox.showinfo(
                "Erfolg",
                f"Dateien erfolgreich gespeichert!\n\n{processed_count} Werte verarbeitet."
                f"\n\nCSV: {save_path}\nExcel: {excel_path}"
            )

    def save_csv_as_excel(self, rows: list[list[str]], excel_path: str) -> None:
        """Speichert CSV-Daten als Excel-Datei. Komplett leere Spalten werden ausgeblendet."""
        if not EXCEL_SUPPORT:
            self.log("HINWEIS: openpyxl nicht installiert - Excel-Export übersprungen.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Hashed Data"

        for row in rows:
            ws.append(row)

        # Komplett leere Spalten ausblenden
        if len(rows) > 1:
            num_cols = max(len(r) for r in rows)
            data_rows = rows[1:]

            for col_idx in range(num_cols):
                col_empty = all(
                    col_idx >= len(r) or not r[col_idx] or str(r[col_idx]).strip() == ""
                    for r in data_rows
                )
                if col_empty:
                    col_letter_xl = get_column_letter(col_idx + 1)
                    ws.column_dimensions[col_letter_xl].hidden = True

        # Spaltenbreite anpassen
        for col in ws.columns:
            col_letter_xl = get_column_letter(col[0].column)
            if ws.column_dimensions[col_letter_xl].hidden:
                continue
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter_xl].width = min(max_length + 2, MAX_COLUMN_WIDTH)

        # Header fett formatieren
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Autofilter setzen
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions

        wb.save(excel_path)


def main() -> None:
    root = tk.Tk()
    app = HashTool(root)
    root.mainloop()


if __name__ == "__main__":
    main()
